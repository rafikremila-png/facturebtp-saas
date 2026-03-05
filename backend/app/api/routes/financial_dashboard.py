"""
Financial Dashboard Routes
API endpoints for financial reports and exports
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, extract
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
from io import BytesIO
import csv
import logging

from app.core.database import get_db
from app.models.models import Invoice, Payment, User
from app.api.deps import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/reports", tags=["Reports"])


class FinancialDashboardResponse(BaseModel):
    """Financial dashboard response schema"""
    total_revenue: float
    total_paid: float
    total_pending: float
    total_overdue: float
    monthly_revenue: List[dict]
    recent_payments: List[dict]
    invoices_by_status: dict
    average_payment_time: Optional[float]
    collection_rate: float


@router.get("/financial", response_model=FinancialDashboardResponse)
async def get_financial_dashboard(
    period: Optional[str] = Query("year", description="Period: month, quarter, year, all"),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Get financial dashboard statistics"""
    user_id = current_user["id"]
    
    # Calculate date range
    now = datetime.now(timezone.utc)
    start_date = None
    
    if period == "month":
        start_date = now - timedelta(days=30)
    elif period == "quarter":
        start_date = now - timedelta(days=90)
    elif period == "year":
        start_date = now - timedelta(days=365)
    # 'all' means no date filter
    
    # Build base query conditions
    conditions = [Invoice.user_id == user_id]
    if start_date:
        conditions.append(Invoice.created_at >= start_date)
    
    # Total revenue (all invoices)
    total_revenue_query = select(func.coalesce(func.sum(Invoice.total_ttc), 0)).where(
        and_(*conditions)
    )
    result = await db.execute(total_revenue_query)
    total_revenue = float(result.scalar() or 0)
    
    # Total paid
    total_paid_query = select(func.coalesce(func.sum(Invoice.amount_paid), 0)).where(
        and_(*conditions)
    )
    result = await db.execute(total_paid_query)
    total_paid = float(result.scalar() or 0)
    
    # Total pending (not paid)
    total_pending = total_revenue - total_paid
    
    # Total overdue (unpaid and past due date)
    overdue_conditions = conditions + [
        Invoice.status.in_(["envoyée", "en_attente", "partiellement_payée"]),
        Invoice.due_date < now
    ]
    overdue_query = select(
        func.coalesce(func.sum(Invoice.total_ttc - Invoice.paid_amount), 0)
    ).where(and_(*overdue_conditions))
    result = await db.execute(overdue_query)
    total_overdue = float(result.scalar() or 0)
    
    # Monthly revenue (last 12 months)
    monthly_revenue = []
    for i in range(12):
        month_start = now.replace(day=1) - timedelta(days=30 * i)
        month_end = (month_start + timedelta(days=32)).replace(day=1)
        
        month_query = select(func.coalesce(func.sum(Invoice.total_ttc), 0)).where(
            and_(
                Invoice.user_id == user_id,
                Invoice.created_at >= month_start,
                Invoice.created_at < month_end
            )
        )
        result = await db.execute(month_query)
        month_total = float(result.scalar() or 0)
        
        month_paid_query = select(func.coalesce(func.sum(Invoice.amount_paid), 0)).where(
            and_(
                Invoice.user_id == user_id,
                Invoice.created_at >= month_start,
                Invoice.created_at < month_end
            )
        )
        result = await db.execute(month_paid_query)
        month_paid = float(result.scalar() or 0)
        
        monthly_revenue.append({
            "month": month_start.strftime("%Y-%m"),
            "month_label": month_start.strftime("%b %Y"),
            "total": month_total,
            "paid": month_paid,
            "pending": month_total - month_paid
        })
    
    monthly_revenue.reverse()  # Oldest first
    
    # Recent payments (from Invoice amount_paid changes)
    recent_payments_query = select(Invoice).where(
        and_(
            Invoice.user_id == user_id,
            Invoice.amount_paid > 0,
            Invoice.status.in_(["payée", "partiellement_payée"])
        )
    ).order_by(Invoice.updated_at.desc()).limit(10)
    
    result = await db.execute(recent_payments_query)
    recent_invoices = result.scalars().all()
    
    recent_payments = []
    for inv in recent_invoices:
        recent_payments.append({
            "id": inv.id,
            "invoice_number": inv.invoice_number,
            "amount": float(inv.amount_paid),
            "total": float(inv.total_ttc),
            "date": inv.updated_at.isoformat() if inv.updated_at else None,
            "status": inv.status
        })
    
    # Invoices by status
    status_counts = {}
    for status_val in ["brouillon", "envoyée", "en_attente", "partiellement_payée", "payée", "annulée"]:
        count_query = select(func.count(Invoice.id)).where(
            and_(Invoice.user_id == user_id, Invoice.status == status_val)
        )
        result = await db.execute(count_query)
        status_counts[status_val] = result.scalar() or 0
    
    # Collection rate (paid / total)
    collection_rate = (total_paid / total_revenue * 100) if total_revenue > 0 else 0
    
    return FinancialDashboardResponse(
        total_revenue=total_revenue,
        total_paid=total_paid,
        total_pending=total_pending,
        total_overdue=total_overdue,
        monthly_revenue=monthly_revenue,
        recent_payments=recent_payments,
        invoices_by_status=status_counts,
        average_payment_time=None,  # Would require payment date tracking
        collection_rate=round(collection_rate, 1)
    )


@router.get("/financial/export/csv")
async def export_financial_csv(
    period: Optional[str] = Query("year"),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Export financial data as CSV"""
    user_id = current_user["id"]
    
    # Get all invoices
    now = datetime.now(timezone.utc)
    conditions = [Invoice.user_id == user_id]
    
    if period == "month":
        conditions.append(Invoice.created_at >= now - timedelta(days=30))
    elif period == "quarter":
        conditions.append(Invoice.created_at >= now - timedelta(days=90))
    elif period == "year":
        conditions.append(Invoice.created_at >= now - timedelta(days=365))
    
    query = select(Invoice).where(and_(*conditions)).order_by(Invoice.created_at.desc())
    result = await db.execute(query)
    invoices = result.scalars().all()
    
    # Create CSV
    output = BytesIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Numéro", "Date", "Client", "Montant HT", "TVA", "Montant TTC",
        "Payé", "Restant", "Statut", "Date d'échéance"
    ])
    
    for inv in invoices:
        writer.writerow([
            inv.invoice_number,
            inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
            inv.title or "",
            round(inv.subtotal_ht or 0, 2),
            round(inv.total_vat or 0, 2),
            round(inv.total_ttc or 0, 2),
            round(inv.amount_paid or 0, 2),
            round((inv.total_ttc or 0) - (inv.amount_paid or 0), 2),
            inv.status or "",
            inv.due_date.strftime("%Y-%m-%d") if inv.due_date else ""
        ])
    
    output.seek(0)
    
    return Response(
        content=output.getvalue().decode('utf-8'),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=factures_{period}_{now.strftime('%Y%m%d')}.csv"
        }
    )


@router.get("/financial/export/excel")
async def export_financial_excel(
    period: Optional[str] = Query("year"),
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Export financial data as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export not available - openpyxl not installed"
        )
    
    user_id = current_user["id"]
    
    # Get all invoices
    now = datetime.now(timezone.utc)
    conditions = [Invoice.user_id == user_id]
    
    if period == "month":
        conditions.append(Invoice.created_at >= now - timedelta(days=30))
    elif period == "quarter":
        conditions.append(Invoice.created_at >= now - timedelta(days=90))
    elif period == "year":
        conditions.append(Invoice.created_at >= now - timedelta(days=365))
    
    query = select(Invoice).where(and_(*conditions)).order_by(Invoice.created_at.desc())
    result = await db.execute(query)
    invoices = result.scalars().all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factures"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    
    # Header
    headers = ["Numéro", "Date", "Client", "Montant HT", "TVA", "Montant TTC",
               "Payé", "Restant", "Statut", "Date d'échéance"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, inv in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=inv.invoice_number)
        ws.cell(row=row, column=2, value=inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "")
        ws.cell(row=row, column=3, value=inv.title or "")
        ws.cell(row=row, column=4, value=round(inv.subtotal_ht or 0, 2))
        ws.cell(row=row, column=5, value=round(inv.total_vat or 0, 2))
        ws.cell(row=row, column=6, value=round(inv.total_ttc or 0, 2))
        ws.cell(row=row, column=7, value=round(inv.amount_paid or 0, 2))
        ws.cell(row=row, column=8, value=round((inv.total_ttc or 0) - (inv.amount_paid or 0), 2))
        ws.cell(row=row, column=9, value=inv.status or "")
        ws.cell(row=row, column=10, value=inv.due_date.strftime("%Y-%m-%d") if inv.due_date else "")
    
    # Adjust column widths
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=factures_{period}_{now.strftime('%Y%m%d')}.xlsx"
        }
    )
